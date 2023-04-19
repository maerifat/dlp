
from __future__ import print_function
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils import quote_sheetname
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import os
import boto3
from botocore.exceptions import ClientError
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
import concurrent.futures
import time
import multiprocessing
import threading
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill


datestamp="April-2023"


def get_document_type(mime_type):
    if mime_type == 'application/vnd.google-apps.document':
        return 'Google Docs'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
        return 'Microsoft Word'
    elif mime_type == 'application/rtf':
        return 'Rich Text Format'
    elif mime_type == 'text/html':
        return 'HTML'
    elif mime_type == 'application/pdf':
        return 'PDF'
    elif mime_type == 'application/x-zip-compressed':
        return "ZIP"
    elif mime_type == 'application/vnd.google-apps.folder':
        return 'Folder'
    elif mime_type == 'image/png':
        return 'PNG'
    elif mime_type == 'text/xml':
        return 'XML'
    elif mime_type == 'application/java-archive':
        return 'JAR'
    elif mime_type == 'application/vnd.google-apps.shortcut':
        return 'SHORTCUT'
    elif mime_type.startswith("video/"):
        return "VIDEO"
    elif mime_type == 'application/epub+zip':
        return 'EPUB'
    elif mime_type == 'application/zip':
        return 'ZIP'
    elif mime_type == 'application/vnd.google-apps.spreadsheet':
        return 'Google Sheets'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        return 'Microsoft Excel'
    elif mime_type == 'application/vnd.google-apps.presentation':
        return 'Google Slides'
    elif mime_type == 'application/vnd.openxmlformats-officedocument.presentationml.presentation':
        return 'Microsoft PowerPoint'
    elif mime_type == 'text/plain':
        return 'Plain text'
    elif mime_type == 'application/vnd.oasis.opendocument.text':
        return 'OpenDocument Text'
    elif mime_type == 'application/vnd.oasis.opendocument.spreadsheet':
        return 'OpenDocument Spreadsheet'
    elif mime_type == 'application/vnd.oasis.opendocument.presentation':
        return 'OpenDocument Presentation'
    elif mime_type.startswith("image/"):
        return "Image"
    else:
        return mime_type








def search_file(user_email):
    """Search file in drive location"""
    # Path to the service account key file downloaded manually
    key_path = "rahul.json"

    # Load credentials from the key file
    creds = None
    if os.path.exists(key_path):
        creds = service_account.Credentials.from_service_account_file(
            key_path, scopes=["https://www.googleapis.com/auth/drive"],
            subject=user_email
        )

    try:
        # create drive api client
        service = build('drive', 'v3', credentials=creds)
        folders = []
        files_data = []

        # Check if file already exists, if it does, load workbook, otherwise create a new one

        wb = Workbook()
        ws = wb.active
        ws.title = "Shared Files"
        ws.append(["File Name", "File Type", "Size", "Owners", "Shared With", "Anonymous Access" ,"View Link", "Created Time", "Modified Time", "Last Viewed Time"])

        page_token = None
        while True:
            # pylint: disable=maybe-no-member
            response = service.files().list(q="'me' in owners",
                                spaces='drive',
                                fields='nextPageToken, files(id, name, mimeType, size, owners, webViewLink, permissions, createdTime, modifiedTime, viewedByMeTime)',
                                pageToken=page_token).execute()
            time.sleep(2)
            for file in response.get('files', []):
                # Process file/folder
                owners = file.get('owners', [])
                owner_emails = [owner.get('emailAddress') for owner in owners]
                shared_with = []
                permissions = file.get('permissions', [])
               
                #print(file.get("name"))
                
        
                for permission in permissions:

                    #print(permission)
                    if ('emailAddress' in permission) and (len(permission['emailAddress']) > 2) :
                        if not permission['emailAddress'].endswith(("@rupifi.com","@rupifi.in","@rupifi.co")):
                            shared_with.append(permission['emailAddress'].strip())
                            AnonymousStatus="No"
                    if 'anyone' in permission['type']:
                        #print(file.get("name"))
                        #print("has excessive permissions") 
                        shared_with.append("anyonewithLink".strip())
                        AnonymousStatus="Yes"
                    
                            
                
                # print(F'Found file: {file.get("name")}, {file.get("id")}, owned by: {", ".join(owner_emails)}, shared with: {",".join(shared_with)}, view link: {file.get("webViewLink")}')

                if shared_with:
                    try:
                        size = int(str(file.get("size")))
                        
                    
                        if size < 1024:
                            size_str = "{} bytes".format(size)
                        elif size < 1024**2:
                            size_str = "{:.0f} KB".format(size/1024)
                        elif size < 1024**3:
                            size_str = "{:.0f} MB".format(size/1024**2)
                        else:
                            size_str = "{:.0f} GB".format(size/1024**3)
                    except :
                        size_str = "Unknown"

                    mime_type= file.get("mimeType")


                    files_data.append((file.get("name"), get_document_type(mime_type), size_str, ", ".join(owner_emails), " , ".join(shared_with).strip(), AnonymousStatus , file.get("webViewLink"), file.get("createdTime"), file.get("modifiedTime"), file.get("viewedByMeTime")))


            page_token = response.get('nextPageToken', None)
            if page_token is None:
                break

        # Append data to the Excel workbook
        for data in files_data:
            ws.append(data)



        for i in range(2, ws.max_row+1):
            url = ws.cell(row=i, column=7).value
            cell = ws.cell(row=i, column=1)
            cell.font = Font(underline="single", color="0563C1")
            cell.hyperlink = url




        font = Font(color='FFFFFF', bold=True)
        fill = PatternFill(start_color='3B5998', end_color='3B5998', fill_type='solid')

        # Set header styles for each column
        for cell in ws[1]:
            cell.font = font
            cell.fill = fill




        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col[0].column in [3]:
                ws.column_dimensions[col_letter].width = 15
            elif col[0].column in [6]:
                ws.column_dimensions[col_letter].width = 18
            elif col[0].column in [1]:
                ws.column_dimensions[col_letter].width = 40
            elif col[0].column in [7]:
                ws.column_dimensions[col_letter].width = 50
            elif col[0].column in [2]:
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 30


        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        white_font = Font(color="FFFFFF")

        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                if cell.value == "Yes":
                    cell.fill = red_fill
                    cell.font = white_font
        # Save Excel workbook
        row_count = ws.max_row
        if row_count >1:
            wb.save(f"{user_email}-{datestamp}.xlsx")
            print(f"{row_count} rows found. hence saved")
        else:
            print("Couldn't save as file seems empty")
    


    except HttpError as error:
        print(F'An error occurred: {error}')
        return None





#send_email('Security Alert! - DLP April...',  'maerifat.majeed@rupifi.com', 'security@security.rupifi.com')





# with open('users.txt', 'r') as file:
#   emails = [email.strip() for email in file.readlines()]
#   print(emails)
def upload_file():
    folder_name=f"DLP-{datestamp}"
    # Path to the service account key file
    key_path = "rahul.json"

    # Email address of the user to impersonate
    email = "security@rupifi.com"

    # ID of the folder to upload the file to
    #folder_id = "11FZL0KKbH4hiSmZKkz0usT-qGsEkZGUW"






    # If the folder doesn't exist, create it





    # Google Drive API scope
    SCOPES = ["https://www.googleapis.com/auth/drive"]

    # Authenticate and impersonate the user
    creds = None
    if os.path.exists(key_path):
        creds = service_account.Credentials.from_service_account_file(
            key_path, scopes=SCOPES, subject=email
        )

    # Create a Drive API client
    drive_service = build("drive", "v3", credentials=creds)

    folder_query = "mimeType='application/vnd.google-apps.folder' and trashed=false and name='%s'" % folder_name
    folders = drive_service.files().list(q=folder_query, fields='files(id)').execute().get('files', [])

    if not folders:
        folder_metadata = {
            'name': folder_name,
            'mimeType': 'application/vnd.google-apps.folder'
        }
        folder = drive_service.files().create(body=folder_metadata, fields='id').execute()
        folder_id = folder.get('id')
    else:
        folder_id = folders[0].get('id')



    # Path to the file to upload
    file_path = f"{user_email}-{datestamp}.xlsx"

    # Name to give the file in Google Drive
    file_name = os.path.basename(file_path)

    # Upload the file to the folder
    file_metadata = {"name": file_name, "parents": [folder_id]}
    media = MediaFileUpload(file_path, resumable=True)
    file = drive_service.files().create(body=file_metadata, media_body=media, fields="id, webViewLink").execute()

    # Share the file with a user with view-only access
    
    permission = {"type": "user",
                "role": "reader",
                "emailAddress": user_email,
                'viewersCanCopyContent': False,
                'copyRequiresWriterPermission': True,
                'capabilities': {
                        'canDownload': False,
                        'canPrint': False,
                        'canCopy': False
                    },
                    "expirationTime": (datetime.utcnow() + timedelta(days=7)).isoformat() + "Z"  # set expiration to 7 days from now
                }
    try:
        drive_service.permissions().create(fileId=file["id"], body=permission,sendNotificationEmail=False).execute()
    except HttpError as error:
        print(f"An error occurred: {error}")
        file = None

    # Print the view link of the file

    try:
        if file:
            global view_link
            view_link = file.get("webViewLink")
            print(f"View link: {view_link}")
            if "spreadsheets" in view_link:
                print(f"{view_link} -after use")
                send_email(f'Security Alert! {folder_name}',  user_email)
                del view_link
            else:
                print("Email couldn't be send.")


        else:
            print("File upload and sharing failed.")
    except Exception as e:
        print(e)
        





def send_email(subject,  user_email):
    # Create a Boto3 session and specify the AWS profile to use
    session = boto3.Session(profile_name='sec')
    
    # Create an SES client using the session
    client = session.client('ses', region_name='ap-south-1')

    Name= user_email.split("@")[0].split(".")[0].capitalize()
   
    

    
    RECIPIENT = user_email

    # Create a multipart message
    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = f'Rupifi Security Team <security@security.rupifi.com>'
    msg['To'] = RECIPIENT
    msg['Reply-To'] = "security@rupifi.com"



    html_content = f"""

<html>
<head></head>
<body><font color="black">
  <b>Hi {Name}</b>
  <p>As a valued member of our team, we would like to remind you of our company's strict policy on Data Loss Prevention (DLP) to protect the confidentiality and integrity of our business data.</p>

<p>We recently conducted an audit and found that you have shared some company files outside our domain. We have attached a Google Spreadsheet that contains the names of the files you have shared. Please check the permissions of these files and restrict them accordingly. This will help us ensure that our business data remains secure and is not exposed to unauthorized individuals.<br></br>
We take data security very seriously, and we expect all our team members to comply with our policies.</p>

The attached Google Spreadsheet will expire in seven days, so please take action promptly. If you have any questions or concerns regarding this matter, please feel free to reach out to the Security Team.<br>
<a href="{view_link}">Files you share outside organisation.</a>
<p><i>Tips: Do Not share the files having sensitive information with permissions as anyoneWithLink.<br>
Do Not share sensitive files outside our domain with any of our partners without having expiration property.<br>
Cross Check Folder permissions which are inherited to the files inside.
<i></p>

<p>Thank you for your cooperation in maintaining the security of our business data.<p>
</font>
Best regards,<br>
Security Team<br>
Rupifi<br>
security@rupifi.com






</body>
</html>
"""

    msg_body = MIMEText(html_content, 'html')
    msg.attach(msg_body)    



    
    # Try to send the email
    try:
        # Create an SES client
        ses_client = boto3.client('ses')
        
        # Send the email
        response = ses_client.send_raw_email(
            Source=msg['From'],
            Destinations=[RECIPIENT],
            RawMessage={
                'Data': msg.as_string()
            }
        )
        print("Email sent successfully.")
        print(f"Email sent to {user_email} with Message ID:", response['MessageId'])
        
    except ClientError as e:
        print("Email delivery failed: ", e.response['Error']['Message'])
        









with open('users.txt', 'r') as file:
    emails = [email.strip() for email in file]
    #print(emails)

start_time = time.time()

for user_email in emails:
    print(f"\n\nUser: {user_email}")
    try:
        search_file(user_email)
        if os.path.exists(f"{user_email}-{datestamp}.xlsx"):
            upload_file()
            print("File uploaded on drive- 200")
        else:
            print("Upload Failed - 404")
    except:
        print(Exception)








end_time = time.time()

execution_time =  end_time - start_time 
print(execution_time)
